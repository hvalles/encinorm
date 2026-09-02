"""Renderer Excel (openpyxl, dependencia opcional)."""

from __future__ import annotations

from ..models import Chart, Detail, Group, Pivot
from ._format import excel_number_format

_COLOR_OPS = {
    "lt": lambda a, b: a < b,
    "le": lambda a, b: a <= b,
    "gt": lambda a, b: a > b,
    "ge": lambda a, b: a >= b,
    "eq": lambda a, b: a == b,
    "ne": lambda a, b: a != b,
}


class ExcelRenderer:
    def __init__(self, styles: dict | None = None, formulas: bool = False):
        self.styles = styles or {}
        self.formulas = formulas

    def render(self, result, ws=None, styles: dict | None = None, formulas: bool | None = None):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError as exc:  # pragma: no cover - depende del entorno
            raise ImportError("openpyxl no está instalado; instala el extra `excel`") from exc

        if ws is None:
            wb = Workbook()
            ws = wb.active
        self._ws = ws
        self._result = result
        self._formulas = formulas if formulas is not None else self.formulas
        self._row = 1

        # KPIs
        for kpi in result.kpis:
            ws.cell(self._row, 1, kpi.label)
            cell = ws.cell(self._row, 2, kpi.value)
            nf = excel_number_format(kpi.format)
            if nf:
                cell.number_format = nf
            self._row += 1

        # encabezado de columnas
        for j, col in enumerate(result.columns, start=1):
            ws.cell(self._row, j, col).font = Font(bold=True)
        self._row += 1

        self._walk(result.root)
        return ws

    def _walk(self, node):
        ws = self._ws
        from openpyxl.styles import Font

        if isinstance(node, Group):
            if node.header:
                self._full_row(node.header, bold=True)
            start = None
            if self._formulas:
                start = self._row
            for child in node.children:
                self._walk(child)
            if node.footer:
                self._full_row(node.footer, bold=True)
            for t in node.totals:
                label = t.label or t.name or t.operator
                col_idx = self._column_index(t.column)
                if (
                    self._formulas
                    and t.operator == "sum"
                    and t.expression is None
                    and col_idx is not None
                    and start is not None
                    and self._row > start
                ):
                    value = self._sum_formula(col_idx, start, self._row - 1)
                else:
                    value = t.value
                fmt = t.format or (self._result.formats.get(t.column) if t.column else None)
                self._total_row(label, value, col_idx, fmt)
        elif isinstance(node, Detail):
            for j, col in enumerate(self._result.columns, start=1):
                value = node.row.get(col)
                cell = ws.cell(self._row, j, value)
                nf = excel_number_format(self._result.formats.get(col))
                if nf:
                    cell.number_format = nf
                self._apply_conditional(cell, col, value)
            self._row += 1
        elif isinstance(node, Chart):
            self._chart(node)
        elif isinstance(node, Pivot):
            self._pivot(node)

    def _column_index(self, column):
        if column is None or column not in self._result.columns:
            return None
        return self._result.columns.index(column) + 1

    def _sum_formula(self, col_idx, start, end):
        from openpyxl.utils import get_column_letter

        letter = get_column_letter(col_idx)
        return f"=SUM({letter}{start}:{letter}{end})"

    def _full_row(self, text, bold=False):
        from openpyxl.styles import Font

        cell = self._ws.cell(self._row, 1, text)
        if bold:
            cell.font = Font(bold=True)
        self._row += 1

    def _total_row(self, label, value, col_idx, fmt):
        ws = self._ws
        from openpyxl.styles import Font

        ncols = len(self._result.columns) or 1
        if col_idx is None:
            col_idx = ncols
        ws.cell(self._row, 1, label).font = Font(bold=True)
        cell = ws.cell(self._row, col_idx + 1 if col_idx < ncols else col_idx, value)
        cell.font = Font(bold=True)
        nf = excel_number_format(fmt)
        if nf and not (isinstance(value, str) and value.startswith("=")):
            cell.number_format = nf
        self._row += 1

    def _apply_conditional(self, cell, column, value):
        from openpyxl.styles import Font

        if value is None:
            return
        color = None
        bold = False
        for rule in self._result.styles:
            if rule.column is not None and rule.column != column:
                continue
            op = _COLOR_OPS.get(rule.when)
            if op is not None and op(value, rule.value):
                if rule.style.get("color"):
                    color = rule.style["color"]
                if rule.style.get("bold"):
                    bold = True
        if color or bold:
            cell.font = Font(color=color, bold=bold)

    def _chart(self, node):
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
        from openpyxl.styles import Font

        self._full_row(f"{node.kind} {node.title or ''}", bold=True)
        if not node.series:
            self._row += 0
            return
        # escribir datos en un área auxiliar y añadir gráfico nativo
        data_start = self._row
        self._ws.cell(self._row, 1, "")
        for j, lab in enumerate(node.labels, start=2):
            self._ws.cell(self._row, j, lab).font = Font(bold=True)
        self._row += 1
        for s in node.series:
            self._ws.cell(self._row, 1, s.label or "")
            for j, v in enumerate(s.values, start=2):
                self._ws.cell(self._row, j, v)
            self._row += 1
        data_end = self._row - 1
        ncols = max(len(node.labels), 1) + 1

        chart_cls = {"pie": PieChart, "bar": BarChart, "line": LineChart}.get(node.kind, BarChart)
        chart = chart_cls()
        chart.title = node.title or node.kind
        data = Reference(self._ws, min_col=2, min_row=data_start + 1,
                         max_col=ncols, max_row=data_end)
        chart.add_data(data, titles_from_data=False)
        cats = Reference(self._ws, min_col=2, min_row=data_start, max_col=ncols)
        chart.set_categories(cats)
        self._ws.add_chart(chart, f"A{self._row + 1}")
        self._row += 1

    def _pivot(self, node):
        from openpyxl.styles import Font

        self._full_row(f"{node.title or 'pivot'}", bold=True)
        start_col = 1
        # encabezado de columnas
        self._ws.cell(self._row, start_col, "").font = Font(bold=True)
        for j, c in enumerate(node.columns, start=start_col + 1):
            self._ws.cell(self._row, j, c).font = Font(bold=True)
        self._row += 1
        for i, r in enumerate(node.rows):
            self._ws.cell(self._row, start_col, r)
            for j, v in enumerate(node.cells[i], start=start_col + 1):
                self._ws.cell(self._row, j, v)
            self._row += 1
